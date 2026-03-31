import os
import re
import shutil
import sys

import fitz
import pdfplumber
from PyQt6.QtCore import QPoint, Qt, QUrl
from PyQt6.QtGui import QDesktopServices, QImage, QPainter, QPen, QPixmap
from PyQt6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QFileDialog,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QProgressBar,
    QPushButton,
    QScrollArea,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)


class PDFViewer(QLabel):
    def __init__(self, pdf_path):
        super().__init__()

        self.doc = fitz.open(pdf_path)
        self.page = self.doc[0]

        pix = self.page.get_pixmap()
        img = QImage.fromData(pix.tobytes("png"))
        self.image = QPixmap.fromImage(img)

        self.setPixmap(self.image)

        self.start = QPoint()
        self.end = QPoint()
        self.drawing = False

    def mousePressEvent(self, event):
        self.start = event.pos()
        self.end = event.pos()
        self.drawing = True
        self.update()

    def mouseMoveEvent(self, event):
        if self.drawing:
            self.end = event.pos()
            self.update()

    def mouseReleaseEvent(self, event):
        self.end = event.pos()
        self.drawing = False

        x0 = min(self.start.x(), self.end.x())
        y0 = min(self.start.y(), self.end.y())
        x1 = max(self.start.x(), self.end.x())
        y1 = max(self.start.y(), self.end.y())

        print(f"IZBRAN PRAVOKOTNIK: ({x0}, {y0}, {x1}, {y1})")
        self.update()

    def paintEvent(self, event):
        super().paintEvent(event)

        if self.drawing:
            painter = QPainter(self)
            pen = QPen(Qt.GlobalColor.red, 2)
            painter.setPen(pen)
            painter.drawRect(
                self.start.x(),
                self.start.y(),
                self.end.x() - self.start.x(),
                self.end.y() - self.start.y(),
            )


class PDFSorterApp(QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Razvrscevalnik PDF velikosti")
        self.setGeometry(250, 180, 1080, 780)

        self.source_folder = ""
        self.destination_folder = ""
        self.excel_file_path = ""
        self.group_pdf_source_folder = ""
        self.group_stp_source_folder = ""
        self.group_destination_folder = ""

        self.combined_output_folder = "Izbrisane glave"
        self.unsorted_output_folder = "neuvr\u0161\u010deni"

        self.is_running = False
        self.cancel_requested = False
        self.stats = {"processed": 0, "matched": 0, "unsorted": 0, "errors": 0}

        self.overlays = {
            "A4": [(55, 730, 173, 813)],
            "A3": [(650, 730, 762, 813)],
            "A2": [(1145, 1075, 1260, 1161)],
            "A1": [(1845, 1575, 1961, 1655)],
            "A0": [(2832, 2273, 2940, 2357)],
        }

        self.build_ui()
        self.update_button_states()

    def build_ui(self):
        layout = QVBoxLayout()

        self.source_btn = QPushButton("Izberi izvorno mapo")
        self.source_btn.clicked.connect(self.select_source)
        self.source_label = QLabel("Izvorna mapa: ni izbrana")
        self.source_label.setWordWrap(True)

        source_row = QHBoxLayout()
        source_row.addWidget(self.source_btn, 0)
        source_row.addWidget(self.source_label, 1)
        layout.addLayout(source_row)

        self.dest_btn = QPushButton("Izberi ciljno mapo")
        self.dest_btn.clicked.connect(self.select_destination)
        self.dest_label = QLabel("Ciljna mapa: ni izbrana")
        self.dest_label.setWordWrap(True)

        dest_row = QHBoxLayout()
        dest_row.addWidget(self.dest_btn, 0)
        dest_row.addWidget(self.dest_label, 1)
        layout.addLayout(dest_row)

        self.excel_btn = QPushButton("Izberi Excel (.xlsx) datoteko")
        self.excel_btn.clicked.connect(self.select_excel_file)
        self.excel_label = QLabel("Excel datoteka: ni izbrana")
        self.excel_label.setWordWrap(True)

        excel_row = QHBoxLayout()
        excel_row.addWidget(self.excel_btn, 0)
        excel_row.addWidget(self.excel_label, 1)
        layout.addLayout(excel_row)

        self.start_btn = QPushButton("Zacni sortiranje (PDF)")
        self.start_btn.clicked.connect(self.sort_pdfs)

        self.group_by_excel_btn = QPushButton("Razvrsti po Excelu (PDF + STP)")
        self.group_by_excel_btn.clicked.connect(self.group_files_by_excel)

        self.stop_btn = QPushButton("Ustavi")
        self.stop_btn.clicked.connect(self.stop_sorting)

        self.open_sorted_btn = QPushButton("Odpri mapo obdelanih")
        self.open_sorted_btn.clicked.connect(self.open_processed_folder)

        self.open_unsorted_btn = QPushButton("Odpri mapo neuvrscenih")
        self.open_unsorted_btn.clicked.connect(self.open_unsorted_folder)

        self.inspect_btn = QPushButton("Preglej koordinate PDF")
        self.inspect_btn.clicked.connect(self.open_pdf_viewer)

        action_row = QHBoxLayout()
        action_row.addWidget(self.start_btn)
        action_row.addWidget(self.group_by_excel_btn)
        action_row.addWidget(self.stop_btn)
        action_row.addWidget(self.open_sorted_btn)
        action_row.addWidget(self.open_unsorted_btn)
        action_row.addWidget(self.inspect_btn)
        layout.addLayout(action_row)

        self.progress_label = QLabel("Napredek: 0 / 0")
        self.progress_bar = QProgressBar()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(1)
        self.progress_bar.setValue(0)

        progress_row = QHBoxLayout()
        progress_row.addWidget(self.progress_label, 0)
        progress_row.addWidget(self.progress_bar, 1)
        layout.addLayout(progress_row)

        self.summary_label = QLabel("Pripravljeno.")
        layout.addWidget(self.summary_label)

        self.results_table = QTableWidget(0, 4)
        self.results_table.setHorizontalHeaderLabels(
            ["Datoteka", "Najdena/Ujemajoca vrednost", "Status", "Izhodna pot"]
        )
        header = self.results_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.results_table.verticalHeader().setVisible(False)
        self.results_table.setAlternatingRowColors(True)
        self.results_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        layout.addWidget(self.results_table, 1)

        self.log = QTextEdit()
        self.log.setReadOnly(True)
        self.log.setFixedHeight(155)
        self.log.setStyleSheet(
            """
            QTextEdit {
                font-size: 12px;
            }
            """
        )
        layout.addWidget(self.log)

        self.setLayout(layout)

    def update_button_states(self):
        has_source_and_dest = bool(self.source_folder and self.destination_folder)
        can_start_size_sort = has_source_and_dest and not self.is_running
        can_group_by_excel = bool(self.excel_file_path) and not self.is_running

        self.start_btn.setEnabled(can_start_size_sort)
        self.group_by_excel_btn.setEnabled(can_group_by_excel)
        self.stop_btn.setEnabled(self.is_running)

        self.source_btn.setEnabled(not self.is_running)
        self.dest_btn.setEnabled(not self.is_running)
        self.excel_btn.setEnabled(not self.is_running)
        self.inspect_btn.setEnabled(not self.is_running)

        combined_folder, unsorted_folder = self.get_output_paths()
        self.open_sorted_btn.setEnabled(
            (not self.is_running) and bool(combined_folder) and os.path.isdir(combined_folder)
        )
        self.open_unsorted_btn.setEnabled(
            (not self.is_running) and bool(unsorted_folder) and os.path.isdir(unsorted_folder)
        )

    def get_output_paths(self):
        if not self.destination_folder:
            return "", ""

        combined_folder = os.path.join(self.destination_folder, self.combined_output_folder)
        unsorted_folder = os.path.join(self.destination_folder, self.unsorted_output_folder)
        return combined_folder, unsorted_folder

    def select_source(self):
        folder = QFileDialog.getExistingDirectory(self, "Izberi izvorno mapo")
        if folder:
            self.source_folder = folder
            self.source_label.setText(f"Izvorna mapa: {folder}")
            self.update_button_states()

    def select_destination(self):
        folder = QFileDialog.getExistingDirectory(self, "Izberi ciljno mapo")
        if folder:
            self.destination_folder = folder
            self.dest_label.setText(f"Ciljna mapa: {folder}")
            self.update_button_states()

    def select_excel_file(self):
        file, _ = QFileDialog.getOpenFileName(
            self,
            "Izberi Excel datoteko",
            "",
            "Excel datoteke (*.xlsx)",
        )
        if file:
            self.excel_file_path = file
            self.excel_label.setText(f"Excel datoteka: {file}")
            self.log_message("Excel datoteka je izbrana.")
            self.update_button_states()

    def log_message(self, message):
        self.log.append(message)

    def reset_run_ui(self, total_files):
        self.results_table.setRowCount(0)
        self.progress_bar.setValue(0)
        self.progress_bar.setMaximum(max(1, total_files))
        self.progress_label.setText(f"Napredek: 0 / {total_files}")
        self.stats = {"processed": 0, "matched": 0, "unsorted": 0, "errors": 0}
        self.summary_label.setText("V teku...")

    def stop_sorting(self):
        if self.is_running:
            self.cancel_requested = True
            self.log_message("Ustavitev zahtevana. Dokoncujem trenutno datoteko...")

    def begin_run(self):
        self.is_running = True
        self.cancel_requested = False
        self.update_button_states()

    def finish_run(self, total_files, matched_label="Ujemanja"):
        self.is_running = False
        self.update_button_states()

        summary = (
            f"Obdelano {self.stats['processed']}/{total_files} | "
            f"{matched_label}: {self.stats['matched']} | "
            f"Neuvrsceni: {self.stats['unsorted']} | "
            f"Napake: {self.stats['errors']}"
        )
        self.summary_label.setText(summary)
        self.log_message(summary)

    def sort_pdfs(self):
        if not self.source_folder or not self.destination_folder:
            self.log_message("Najprej izberite izvorno in ciljno mapo.")
            return

        pdf_files = sorted(
            file for file in os.listdir(self.source_folder) if file.lower().endswith(".pdf")
        )
        total_files = len(pdf_files)
        self.reset_run_ui(total_files)

        if total_files == 0:
            self.summary_label.setText("V izvorni mapi ni PDF datotek.")
            self.log_message("V izvorni mapi ni PDF datotek.")
            self.update_button_states()
            return

        combined_folder, unsorted_folder = self.get_output_paths()
        os.makedirs(combined_folder, exist_ok=True)
        os.makedirs(unsorted_folder, exist_ok=True)

        self.begin_run()
        self.log_message(f"Zaceto sortiranje PDF velikosti za {total_files} datotek.")

        for file in pdf_files:
            QApplication.processEvents()
            if self.cancel_requested:
                self.log_message("Sortiranje velikosti je ustavljeno.")
                break

            file_path = os.path.join(self.source_folder, file)
            detected_size = "-"
            status = ""
            output_path = ""

            try:
                with pdfplumber.open(file_path) as pdf:
                    detected_size = self.detect_sheet_size(pdf)

                if detected_size:
                    output_path = self.get_unique_output_path(combined_folder, file)
                    self.overlay_and_save_pdf(file_path, output_path, detected_size)
                    status = "Ujemanje (velikost)"
                    self.stats["matched"] += 1
                    self.log_message(f"Obdelano: {file} ({detected_size})")
                else:
                    output_path = self.copy_to_unsorted(file_path, file, unsorted_folder)
                    status = "Neuvrsceno (velikost ni najdena)"
                    self.stats["unsorted"] += 1
                    self.log_message(f"Velikost ni najdena: {file}. Kopirano v neuvrscene.")

            except Exception as e:
                output_path = self.copy_to_unsorted(file_path, file, unsorted_folder)
                status = "Napaka -> neuvrsceno"
                self.stats["errors"] += 1
                self.log_message(f"Napaka pri branju datoteke {file}: {e}. Kopirano v neuvrscene.")

            self.stats["processed"] += 1
            self.add_result_row(file, detected_size, status, output_path)
            self.update_progress(self.stats["processed"], total_files)

        self.finish_run(total_files, matched_label="Ujemanja")

    def group_files_by_excel(self):
        if not self.excel_file_path:
            self.log_message("Najprej izberite Excel datoteko.")
            return

        pdf_source_folder = QFileDialog.getExistingDirectory(
            self,
            "Izberi mapo PDF datotek (Preklic za preskok PDF datotek)",
            self.group_pdf_source_folder or self.source_folder,
        )
        stp_source_folder = QFileDialog.getExistingDirectory(
            self,
            "Izberi mapo STP datotek (Preklic za preskok STP datotek)",
            self.group_stp_source_folder or self.source_folder,
        )
        group_destination_folder = QFileDialog.getExistingDirectory(
            self,
            "Izberi ciljno mapo za razvrscanje po Excelu",
            self.group_destination_folder or self.destination_folder,
        )

        if not group_destination_folder:
            self.log_message("Razvrscanje po Excelu prekinjeno: ciljna mapa ni izbrana.")
            return

        if not pdf_source_folder and not stp_source_folder:
            self.log_message("Razvrscanje po Excelu prekinjeno: ni izbrane mape PDF ali STP.")
            return

        self.group_pdf_source_folder = pdf_source_folder
        self.group_stp_source_folder = stp_source_folder
        self.group_destination_folder = group_destination_folder

        self.log_message(
            f"Izbrane mape -> PDF: '{pdf_source_folder or '-'}', "
            f"STP: '{stp_source_folder or '-'}', "
            f"Cilj: '{group_destination_folder}'"
        )

        if not self.destination_folder:
            self.destination_folder = group_destination_folder
            self.dest_label.setText(f"Ciljna mapa: {group_destination_folder}")
            self.update_button_states()

        files_to_process = []
        if pdf_source_folder:
            files_to_process.extend(
                self.collect_files_with_extension(pdf_source_folder, ".pdf")
            )
        if stp_source_folder:
            files_to_process.extend(
                self.collect_files_with_extension(stp_source_folder, ".stp")
            )

        total_files = len(files_to_process)
        self.reset_run_ui(total_files)

        if total_files == 0:
            self.summary_label.setText("V izbranih mapah ni .pdf ali .stp datotek.")
            self.log_message("V izbranih mapah ni .pdf ali .stp datotek.")
            self.update_button_states()
            return

        try:
            drawing_records, rows_loaded = self.load_excel_mapping(self.excel_file_path)
        except Exception as e:
            self.summary_label.setText("Branje Excel datoteke ni uspelo.")
            self.log_message(f"Napaka pri branju Excel datoteke: {e}")
            self.update_button_states()
            return

        if not drawing_records:
            self.summary_label.setText("V Excelu ni veljavnih vrstic.")
            self.log_message(
                "Ni veljavnih vrstic z vrednostma v stolpcih 'Drawing no.' in 'KOOPERANT/kooperacija'."
            )
            self.update_button_states()
            return

        match_index = self.build_drawing_match_index(drawing_records)

        unsorted_folder = os.path.join(group_destination_folder, self.unsorted_output_folder)
        os.makedirs(unsorted_folder, exist_ok=True)

        self.begin_run()
        self.log_message(
            f"Zaceto razvrscanje po Excelu za {total_files} datotek. "
            f"Nalozenih vrstic za ujemanje: {rows_loaded}."
        )

        for file_path, file in files_to_process:
            QApplication.processEvents()
            if self.cancel_requested:
                self.log_message("Razvrscanje po Excelu je ustavljeno.")
                break

            base_name = os.path.splitext(file)[0]
            match_value = "-"
            status = ""
            output_path = ""

            matched_kooperants, matched_drawings = self.match_file_to_kooperants(
                base_name,
                match_index,
            )

            if matched_kooperants:
                copied_paths = []
                try:
                    for kooperant_name in sorted(matched_kooperants):
                        target_folder = os.path.join(
                            group_destination_folder, self.sanitize_folder_name(kooperant_name)
                        )
                        os.makedirs(target_folder, exist_ok=True)

                        matched_output_path = self.get_unique_output_path(target_folder, file)
                        shutil.copy2(file_path, matched_output_path)
                        copied_paths.append(matched_output_path)

                    match_value = ", ".join(sorted(matched_drawings))
                    status = f"Razvrsceno (Excel -> {len(matched_kooperants)} KOOPERANT)"
                    output_path = " | ".join(copied_paths)
                    self.stats["matched"] += 1
                    self.log_message(
                        f"Razvrsceno: {file} | Ujemanje Drawing no.: {match_value} | "
                        f"KOOPERANT: {', '.join(sorted(matched_kooperants))}"
                    )
                except Exception as e:
                    output_path = self.copy_to_unsorted(file_path, file, unsorted_folder)
                    status = "Napaka pri kopiranju -> neuvrsceno"
                    self.stats["errors"] += 1
                    self.log_message(f"Napaka pri kopiranju datoteke {file}: {e}. Kopirano v neuvrscene.")
            else:
                output_path = self.copy_to_unsorted(file_path, file, unsorted_folder)
                status = "Neuvrsceno (brez ujemanja Drawing no. v imenu datoteke)"
                self.stats["unsorted"] += 1
                self.log_message(f"Brez ujemanja Drawing no.: {file}. Kopirano v neuvrscene.")

            self.stats["processed"] += 1
            self.add_result_row(file, match_value, status, output_path)
            self.update_progress(self.stats["processed"], total_files)

        self.finish_run(total_files, matched_label="Razvrsceno")

    def collect_files_with_extension(self, folder_path, extension):
        files = []
        try:
            for root, _, filenames in os.walk(folder_path):
                for file in sorted(filenames):
                    if file.lower().endswith(extension):
                        files.append((os.path.join(root, file), file))
        except Exception as e:
            self.log_message(f"Mape ni mogoce prebrati '{folder_path}': {e}")
        return files

    def load_excel_mapping(self, excel_path):
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise RuntimeError(
                "Paket openpyxl je obvezen. Namestite ga z ukazom: pip install openpyxl"
            ) from e

        workbook = load_workbook(excel_path, data_only=True, read_only=True)
        drawing_records = []
        loaded_rows = 0

        for sheet in workbook.worksheets:
            drawing_col_idx, kooperant_col_idx, header_row = self.find_header_columns(sheet)
            if drawing_col_idx is None or kooperant_col_idx is None:
                self.log_message(
                    "List je preskocen "
                    f"'{sheet.title}' (stolpca 'Drawing no.' in 'KOOPERANT/kooperacija' nista bila najdena)."
                )
                continue

            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                drawing_val = row[drawing_col_idx] if drawing_col_idx < len(row) else None
                kooperant_val = row[kooperant_col_idx] if kooperant_col_idx < len(row) else None

                if drawing_val is None or kooperant_val is None:
                    continue

                drawing_text = str(drawing_val).strip()
                kooperant_text = str(kooperant_val).strip()
                if not drawing_text or not kooperant_text:
                    continue

                drawing_key = self.normalize_text_key(drawing_text)
                numeric_key = self.normalize_numeric_key(drawing_text)
                if not drawing_key and not numeric_key:
                    continue

                drawing_records.append(
                    {
                        "drawing_raw": drawing_text,
                        "drawing_key": drawing_key,
                        "drawing_numeric": numeric_key,
                        "kooperant": kooperant_text,
                    }
                )
                loaded_rows += 1

        workbook.close()
        return drawing_records, loaded_rows

    def find_header_columns(self, sheet):
        max_scan_rows = min(25, sheet.max_row)
        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True),
            start=1,
        ):
            drawing_idx = None
            kooperant_idx = None

            for col_idx, value in enumerate(row):
                normalized_header = self.normalize_header(value)
                if normalized_header.startswith("drawingno"):
                    drawing_idx = col_idx
                elif (
                    normalized_header.startswith("kooperacija")
                    or normalized_header.startswith("kooperant")
                ):
                    kooperant_idx = col_idx

            if drawing_idx is not None and kooperant_idx is not None:
                return drawing_idx, kooperant_idx, row_idx

        return None, None, None

    def build_drawing_match_index(self, drawing_records):
        index = {
            "text_entries": [],
            "numeric_entries": {},
        }

        for record in drawing_records:
            drawing_raw = record["drawing_raw"]
            drawing_key = record["drawing_key"]
            drawing_numeric = record["drawing_numeric"]
            kooperant = record["kooperant"]

            if drawing_key:
                index["text_entries"].append((drawing_key, kooperant, drawing_raw))

            if drawing_numeric:
                if drawing_numeric not in index["numeric_entries"]:
                    index["numeric_entries"][drawing_numeric] = []
                index["numeric_entries"][drawing_numeric].append((kooperant, drawing_raw))

        return index

    def match_file_to_kooperants(self, file_base_name, match_index):
        matched_kooperants = set()
        matched_drawings = set()

        file_key = self.normalize_text_key(file_base_name)
        file_numeric_keys = self.extract_numeric_keys(file_base_name)

        # Numeric matching is primary: handles leading-zero filenames such as 0000094682_prt.stp.
        for numeric_key in file_numeric_keys:
            if numeric_key in match_index["numeric_entries"]:
                for kooperant, drawing_raw in match_index["numeric_entries"][numeric_key]:
                    matched_kooperants.add(kooperant)
                    matched_drawings.add(drawing_raw)

        # Fallback textual "contains" matching for non-numeric drawing identifiers.
        if not matched_kooperants:
            for drawing_key, kooperant, drawing_raw in match_index["text_entries"]:
                if drawing_key and drawing_key in file_key:
                    matched_kooperants.add(kooperant)
                    matched_drawings.add(drawing_raw)

        return matched_kooperants, matched_drawings

    def normalize_header(self, value):
        if value is None:
            return ""
        text = str(value).strip().lower()
        return re.sub(r"[\s._-]+", "", text)

    def normalize_text_key(self, value):
        text = str(value).strip().casefold()
        return re.sub(r"[\s_-]+", "", text)

    def normalize_numeric_key(self, value):
        digits = re.sub(r"\D+", "", str(value))
        if not digits:
            return ""
        return digits.lstrip("0") or "0"

    def extract_numeric_keys(self, value):
        keys = set()
        for token in re.findall(r"\d+", str(value)):
            normalized = token.lstrip("0") or "0"
            keys.add(normalized)
        return keys

    def sanitize_folder_name(self, folder_name):
        safe = re.sub(r'[<>:"/\\|?*]', "_", str(folder_name).strip())
        safe = safe.rstrip(". ")
        return safe if safe else "neznano_kooperacija"

    def detect_sheet_size(self, pdf):
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue

            match = re.search(r"\bA\s*[-]?\s*([0-4])\b", text, re.IGNORECASE)
            if match:
                return f"A{match.group(1)}"

        return None

    def update_progress(self, current, total):
        self.progress_bar.setMaximum(max(1, total))
        self.progress_bar.setValue(current)
        self.progress_label.setText(f"Napredek: {current} / {total}")
        QApplication.processEvents()

    def add_result_row(self, file_name, detected_value, status, output_path):
        row = self.results_table.rowCount()
        self.results_table.insertRow(row)
        self.results_table.setItem(row, 0, QTableWidgetItem(file_name))
        self.results_table.setItem(row, 1, QTableWidgetItem(detected_value))
        self.results_table.setItem(row, 2, QTableWidgetItem(status))
        self.results_table.setItem(row, 3, QTableWidgetItem(output_path))

    def open_folder(self, folder_path):
        if not folder_path or not os.path.isdir(folder_path):
            self.log_message("Mapa ne obstaja.")
            return

        QDesktopServices.openUrl(QUrl.fromLocalFile(folder_path))

    def open_processed_folder(self):
        combined_folder, _ = self.get_output_paths()
        self.open_folder(combined_folder)

    def open_unsorted_folder(self):
        _, unsorted_folder = self.get_output_paths()
        self.open_folder(unsorted_folder)

    def open_pdf_viewer(self):
        file, _ = QFileDialog.getOpenFileName(self, "Izberi PDF", "", "PDF datoteke (*.pdf)")
        if file:
            self.viewer = PDFViewer(file)

            self.scroll = QScrollArea()
            self.scroll.setWidget(self.viewer)
            self.scroll.setWidgetResizable(True)
            self.scroll.setWindowTitle("Izberi obmocje (klik + povleci)")
            self.scroll.showMaximized()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key.Key_Escape:
            self.close()

    def overlay_and_save_pdf(self, input_path, output_path, sheet_size):
        doc = fitz.open(input_path)
        rectangles = self.overlays.get(sheet_size, [])

        for page in doc:
            for coords in rectangles:
                rect = fitz.Rect(coords)
                page.draw_rect(rect, color=(1, 1, 1), fill=(1, 1, 1))

        doc.save(output_path)
        doc.close()

    def copy_to_unsorted(self, file_path, filename, unsorted_folder):
        unsorted_output_path = self.get_unique_output_path(unsorted_folder, filename)
        shutil.copy2(file_path, unsorted_output_path)
        return unsorted_output_path

    def get_unique_output_path(self, target_folder, filename):
        base, ext = os.path.splitext(filename)
        candidate_path = os.path.join(target_folder, filename)
        counter = 1

        while os.path.exists(candidate_path):
            candidate_path = os.path.join(target_folder, f"{base}_{counter}{ext}")
            counter += 1

        return candidate_path


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = PDFSorterApp()
    window.show()
    sys.exit(app.exec())

